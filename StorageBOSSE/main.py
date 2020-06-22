import os
import pandas as pd
from StorageBOSSE.excelio.create_master_input_dict import XlsxReader
from LandBOSSE.landbosse.excelio.XlsxDataframeCache import XlsxDataframeCache
from StorageBOSSE.model.Manager import Manager
import xlsxwriter
from openpyxl import load_workbook

def run_storagebosse(input_dictionary):
    input_output_path = os.path.dirname(__file__)
    # StorageBOSSE uses LandBOSSE's Excel I/O library for reading in data from Excel
    # files. Accordingly, the environment variables used in StorageBOSSE are called
    # LANDBOSSE_INPUT_DIR & LANDBOSSE_OUTPUT_DIR:
    os.environ["LANDBOSSE_INPUT_DIR"] = input_output_path
    os.environ["LANDBOSSE_OUTPUT_DIR"] = input_output_path

    project_data = read_data(input_dictionary['project_list'])
    xlsx_reader = XlsxReader()
    for _, project_parameters in project_data.iterrows():
        project_data_basename = project_parameters['Project data file']

        project_data_sheets = \
            XlsxDataframeCache.read_all_sheets_from_xlsx(project_data_basename)

        # make sure you call create_master_input_dictionary() as soon as
        # labor_cost_multiplier's value is changed.
        master_input_dict = xlsx_reader.create_master_input_dictionary(
                                            project_data_sheets, project_parameters)

        master_input_dict['error'] = dict()

    # master_input_dict = dict()
    output_dict = dict()

    for key, _ in input_dictionary.items():
        master_input_dict[key] = input_dictionary[key]

    # Manager class (1) manages the distribution of inout data for all modules
    # and (2) executes landbosse
    mc = Manager(input_dict=master_input_dict, output_dict=output_dict)
    mc.execute_storagebosse()

    # results dictionary that gets returned by this function:
    results = dict()
    # results['errors'] = []  # TODO uncomment after testing
    if master_input_dict['error']:
        for key, value in master_input_dict['error'].items():
            msg = "Error in " + key + ": " + str(value)
            results['errors'].append(msg)
    else:   # if project runs successfully, return a dictionary with results
        # that are 3 layers deep (but 1-D)
        results['Name'] = 'area undefined'
        # results['Name'] = str(master_input_dict['system_size_MW_DC'])+'MW_'+str(master_input_dict['system_size_MWh'])+'MWh'
        # results['labor_cost_multiplier'] = master_input_dict['labor_cost_multiplier']
        # results['material_cost_multiplier'] = master_input_dict['material_cost_multiplier']
        results['system_size_MW_DC'] = master_input_dict['system_size_MW_DC']
        results['system_size_MWh'] = master_input_dict['system_size_MWh']
        results['total_cost'] = output_dict['total_cost']
        results['total_container_cost'] = output_dict['total_container_cost']
        results['total_bos_cost'] = output_dict['total_bos_cost']
        results['total_bos_cost_before_mgmt'] = output_dict['total_bos_cost_before_mgmt']
        results['total_road_cost'] = output_dict['total_road_cost']
        results['substation_cost'] = output_dict['total_substation_cost']
        results['total_transdist_cost'] = output_dict['total_transdist_cost']
        results['total_foundation_cost'] = output_dict['total_foundation_cost']
        results['total_erection_cost'] = output_dict['total_erection_cost']
        results['total_collection_cost'] = output_dict['total_collection_cost']
        results['total_management_cost'] = output_dict['total_management_cost']


        # L1: assume share substation, transdist, mgmt costs
        results['%_bos_cost'] = results['total_bos_cost']/results['total_cost']*100
        results['shared_cost_L1'] = results['substation_cost'] + results['total_transdist_cost'] + results['total_management_cost']
        results['%_shared_cost_L1'] = results['shared_cost_L1']/results['total_bos_cost']*100
        # L2: L1 + share road/site prep cost
        results['shared_cost_L2'] = results['shared_cost_L1'] + results['total_road_cost']
        results['%_shared_cost_L2'] = results['shared_cost_L2']/results['total_bos_cost']*100
        row = output_dict['total_erection_cost_df'][output_dict['total_erection_cost_df']['Type of cost'] == 'Mobilization']
        # L3: L2 + share crane cost
        results['shared_cost_L3'] = results['shared_cost_L2'] + row.loc[0, 'Cost USD']
        results['%_shared_cost_L3'] = results['shared_cost_L3']/results['total_bos_cost']*100

    return results, output_dict


# This method reads in the two input Excel files (project_list; project_1)
# and stores them as data frames. This method is called internally in
# run_landbosse(), where the data read in is converted to a master input
# dictionary.
def read_data(file_name):
    path_to_project_list = os.path.dirname(__file__)
    sheets = XlsxDataframeCache.read_all_sheets_from_xlsx(file_name,
                                                          path_to_project_list)

    # If there is one sheet, make an empty data frame as a placeholder.
    if len(sheets.values()) == 1:
        first_sheet = list(sheets.values())[0]
        project_list = first_sheet

    # If the parametric and project lists exist, read them
    elif 'Project list' in sheets.keys():
        project_list = sheets['Project list']

    # Otherwise, raise an exception
    else:
        raise KeyError(
            "Project list needs to have a single sheet or sheets named "
            "'Project list' and 'Parametric list'."
        )

    return project_list


def read_weather_data(file_path):
    weather_data = pd.read_csv(file_path,
                               sep=",",
                               header=None,
                               skiprows=5,
                               usecols=[0, 1, 2, 3, 4]
                               )
    return weather_data


class Error(Exception):
    """
        Base class for other exceptions
    """
    pass


class NegativeInputError(Error):
    """
        User entered a negative input. This is an invalid entry.
    """
    pass

# <><><><><><><><> EXAMPLE OF RUNNING THIS StorageBOSSE API <><><><><><><><><><><>
# TODO: uncomment these lines to run StorageBOSSE as a standalone model.
# overwrite keys in input_dict. Can also modify keys in project_list excel file.
# optional key: site_prep_area_m2: defined project area for site preparation. If unspecified, site prep area is
# calculated based on number of containers and road length

##### loop through various project sizes. Find % shared #####

energies = [1, 5, 10, 50, 100, 150, 20, 20, 20, 20, 20, 20]  # MWh
powers = [20, 20, 20, 20, 20, 20, 1, 5, 10, 50, 100, 150]  # MW

for i in range(0, len(energies)):
    input_dict = dict()
    BOS_results = dict()
    BOS_results.update({str(powers[i])+' MW, '+str(energies[i])+'MWh scenario': ' '})
    input_dict['system_size_MW_DC'] = powers[i]
    input_dict['system_size_MWh'] = energies[i]

    if max(energies[i], powers[i]) > 50:
        input_dict['construction_time_months'] = 24
    elif max(energies[i], powers[i]) <= 20:
        input_dict['construction_time_months'] = 12

    elif max(energies[i], powers[i]) <= 10:
        input_dict['construction_time_months'] = 6

# ##### loop through various material multipliers ##########
# energies = 100
# powers = 100
# input_dict = dict()
# BOS_results = dict()
# mcx = [0, 0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5]
# for i in mcx:
#     input_dict['material_cost_multiplier'] = i

# ###### loop through various management multipliers #########
# energies = 100
# powers = 100
# input_dict = dict()
# BOS_results = dict()
# mcx = [0, 0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5]
# for i in mcx:
#     input_dict['management_cost_multiplier'] = i

    ####################################### KEEP (in loop) ####################################
    # input_dict['site_prep_area_m2'] = 100*max(max(energies),max(powers)) + 100

    input_dict['project_list'] = 'project_list_test'  # THE ESSENTIAL INPUT

    BOS_results, detailed_results = run_storagebosse(input_dict)  # ALL I NEEEEED

    """ OUTPUT TO EXCEL FOR TESTING"""
    headers = BOS_results.keys()
    outfile = 'shared_cost_outputs_MW.xlsx'
    # create excel file if it does not exist
    if not os.path.isfile(outfile):
        book = xlsxwriter.Workbook(outfile)
        sheet = book.add_worksheet("TestSheet")
        for (idx, header) in enumerate(headers):
            sheet.write(0, idx, header)
        book.close()
    # open the file
    with open(outfile, 'a+') as xl_file:
        book = load_workbook(outfile)
        sheet = book.get_sheet_by_name('TestSheet')

    values = [BOS_results[key] for key in headers]
    sheet.append(values)
    book.save(filename=outfile)


# <><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><><
